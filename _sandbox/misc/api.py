from flask import Flask, request, jsonify, send_file
from pathlib import Path
import tempfile
import pandas as pd
from io import StringIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook

app = Flask(__name__)

@app.route('/api/json-to-excel', methods=['POST'])
def convert_json_to_excel():
    try:
        data = request.get_json()
        
        wb = Workbook()
        ws = wb.active
        
        if data and len(data) > 0:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                add_cell(
                    wb,
                    row=1,
                    column=col,
                    value=header,
                    font_props={'bold': True},
                    fill_props={'fill_type': 'solid', 'fgColor': 'FF0000'}
                )
            
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, key in enumerate(headers, 1):
                    add_cell(
                        wb,
                        row=row_idx,
                        column=col_idx,
                        value=row_data.get(key, ''),
                        font_props={'bold': True},
                        fill_props={'fill_type': 'solid', 'fgColor': 'FF0000'}
                    )
        
        temp_dir = Path(tempfile.gettempdir())
        file_path = temp_dir / "output.xlsx"
        
        wb.save(file_path)
        
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='json_to_excel.xlsx'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/csv-to-excel', methods=['POST'])
def convert_csv_to_excel():
    try:
        if request.get_data(as_text=True) and ',' in request.get_data(as_text=True) and '\n' in request.get_data(as_text=True):
            df = pd.read_csv(StringIO(request.get_data(as_text=True)))
        elif request.files:
            if 'file' not in request.files:
                return jsonify({"error": "No file provided"}), 400
                
            file = request.files['file']
            
            if file.filename == '' or not file.filename.endswith('.csv'):
                return jsonify({"error": "Invalid file. Please upload a CSV file"}), 400
                
            df = pd.read_csv(file)
        else:
            return jsonify({"error": "No valid CSV data provided. Send either raw CSV data or a CSV file."}), 400
        
        wb = Workbook()
        ws = wb.active
        
        headers = df.columns.tolist()
        for col, header in enumerate(headers, 1):
            add_cell(
                wb,
                row=1,
                column=col,
                value=header,
                font_props={'bold': True},
                fill_props={'fill_type': 'solid', 'fgColor': 'FF0000'}
            )
        
        for row_idx, row in enumerate(df.values, 2):
            for col_idx, value in enumerate(row, 1):
                add_cell(
                    wb,
                    row=row_idx,
                    column=col_idx,
                    value=value,
                    font_props={'bold': True},
                    fill_props={'fill_type': 'solid', 'fgColor': 'FF0000'}
                )
        
        temp_dir = Path(tempfile.gettempdir())
        file_path = temp_dir / "output.xlsx"
        
        wb.save(file_path)
        
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='csv_to_excel.xlsx'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

def add_cell(
    workbook: Workbook,
    sheet_name: str = None,
    row: int = 1,
    column: int = 1,
    value: any = None,
    font_props: dict = None,
    fill_props: dict = None,
    alignment_props: dict = None,
    border_props: dict = None,
    merge_cells: tuple = None
) -> None:

    ws = workbook[sheet_name] if sheet_name else workbook.active
    
    cell = ws.cell(row=row, column=column, value=value)
    
    if font_props:
        cell.font = Font(**font_props)
    
    if fill_props:
        cell.fill = PatternFill(**fill_props)
    
    if alignment_props:
        cell.alignment = Alignment(**alignment_props)
    
    if border_props:
        border = Border(**{
            side: Side(**border_props)
            for side in ['left', 'right', 'top', 'bottom']
        })
        cell.border = border
    
    if merge_cells:
        start_row, start_col, end_row, end_col = merge_cells
        ws.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col
        )
    
    return cell

if __name__ == '__main__':
    app.run(debug=True)